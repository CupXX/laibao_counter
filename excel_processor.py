"""
Excelæ–‡ä»¶å¤„ç†æ¨¡å—
è´Ÿè´£è¯»å–Excelæ–‡ä»¶å¹¶æå–æ˜µç§°æ•°æ®
"""
import pandas as pd
from typing import List, Tuple, Optional, Dict
import re
from openpyxl import load_workbook
import io
import streamlit as st


class ExcelProcessor:
    # å¸¸è§çš„æ˜µç§°/å§“ååˆ—åå…³é”®è¯
    NICKNAME_KEYWORDS = [
        'æ˜µç§°', 'å§“å'
    ]
    
    # å¸¸è§çš„æ—¶é—´åˆ—åå…³é”®è¯
    TIME_KEYWORDS = [
        'æäº¤æ—¶é—´', 'æ—¶é—´', 'æ‰“å¡æ—¶é—´', 'å‚ä¸æ—¶é—´', 'ä¸Šä¼ æ—¶é—´', 
        'time', 'submit_time', 'timestamp', 'æ—¥æœŸæ—¶é—´'
    ]
    
    def __init__(self):
        pass
    
    def find_nickname_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        è‡ªåŠ¨æŸ¥æ‰¾æ˜µç§°åˆ—
        
        Args:
            df: pandas DataFrame
            
        Returns:
            æ˜µç§°åˆ—åï¼Œå¦‚æœæ‰¾ä¸åˆ°è¿”å›None
        """
        columns = df.columns.tolist()
        
        # ç²¾ç¡®åŒ¹é…
        for keyword in self.NICKNAME_KEYWORDS:
            for col in columns:
                if keyword == str(col).strip():
                    return col
        
        # æ¨¡ç³ŠåŒ¹é…
        for keyword in self.NICKNAME_KEYWORDS:
            for col in columns:
                if keyword in str(col):
                    return col
        
        # å¦‚æœæ²¡æœ‰æ‰¾åˆ°ï¼Œæ£€æŸ¥ç¬¬ä¸€åˆ—æ˜¯å¦åŒ…å«æ–‡æœ¬æ•°æ®
        if len(columns) > 0:
            first_col = columns[0]
            # æ£€æŸ¥ç¬¬ä¸€åˆ—å‰å‡ è¡Œæ˜¯å¦ä¸»è¦æ˜¯æ–‡æœ¬
            sample_data = df[first_col].dropna().head(10)
            if len(sample_data) > 0:
                text_count = sum(1 for x in sample_data if isinstance(x, str) and len(str(x).strip()) > 0)
                if text_count / len(sample_data) > 0.7:  # 70%ä»¥ä¸Šæ˜¯æ–‡æœ¬
                    return first_col
        
        return None
    
    def find_time_column(self, df: pd.DataFrame) -> Optional[str]:
        """
        è‡ªåŠ¨æŸ¥æ‰¾æ—¶é—´åˆ—
        
        Args:
            df: pandas DataFrame
            
        Returns:
            æ—¶é—´åˆ—åï¼Œå¦‚æœæ‰¾ä¸åˆ°è¿”å›None
        """
        columns = df.columns.tolist()
        
        # ç²¾ç¡®åŒ¹é…
        for keyword in self.TIME_KEYWORDS:
            for col in columns:
                if keyword == str(col).strip():
                    return col
        
        # æ¨¡ç³ŠåŒ¹é…
        for keyword in self.TIME_KEYWORDS:
            for col in columns:
                if keyword in str(col):
                    return col
        
        return None
    
    def clean_nickname(self, nickname: str) -> str:
        """
        æ¸…ç†æ˜µç§°æ•°æ®
        
        Args:
            nickname: åŸå§‹æ˜µç§°
            
        Returns:
            æ¸…ç†åçš„æ˜µç§°
        """
        if pd.isna(nickname) or nickname is None:
            return ""
        
        nickname = str(nickname).strip()
        
        # ç§»é™¤å¸¸è§çš„æ— æ•ˆå­—ç¬¦å’Œæ ‡è®°
        # ç§»é™¤emojiï¼ˆç®€å•å¤„ç†ï¼‰
        nickname = re.sub(r'[^\w\u4e00-\u9fff\u3400-\u4dbf\s]', '', nickname)
        
        # ç§»é™¤å¤šä½™ç©ºæ ¼
        nickname = re.sub(r'\s+', ' ', nickname).strip()
        
        return nickname
    
    def count_images_in_excel(self, file_content, file_name: str, header_row: int = 0) -> Dict[str, int]:
        """
        ç»Ÿè®¡Excelæ–‡ä»¶ä¸­æ¯ä¸ªæ˜µç§°å¯¹åº”çš„å›¾ç‰‡æ•°é‡ï¼ˆç æ•°ï¼‰
        
        Args:
            file_content: æ–‡ä»¶å†…å®¹ï¼ˆbytesæˆ–file-likeå¯¹è±¡ï¼‰
            file_name: æ–‡ä»¶å
            header_row: åˆ—åæ‰€åœ¨è¡Œï¼ˆ0-basedç´¢å¼•ï¼‰
            
        Returns:
            å­—å…¸ {æ˜µç§°: å›¾ç‰‡æ•°é‡}
        """
        try:
            # é‡ç½®æ–‡ä»¶æŒ‡é’ˆ
            if hasattr(file_content, 'seek'):
                file_content.seek(0)
            
            # å¦‚æœæ˜¯file-likeå¯¹è±¡ï¼Œéœ€è¦è¯»å–åˆ°BytesIO
            if hasattr(file_content, 'read'):
                content = file_content.read()
                if hasattr(file_content, 'seek'):
                    file_content.seek(0)  # é‡ç½®ï¼Œä¾›åç»­ä½¿ç”¨
                file_obj = io.BytesIO(content)
            else:
                file_obj = io.BytesIO(file_content)
            
            # ä½¿ç”¨openpyxlè¯»å–ï¼ˆæ”¯æŒè¶…é“¾æ¥ï¼‰
            wb = load_workbook(file_obj, data_only=False)
            ws = wb.active
            
            # è·å–åˆ—åï¼ˆheader_rowæ˜¯0-basedï¼Œä½†openpyxlä½¿ç”¨1-basedï¼‰
            openpyxl_header_row = header_row + 1
            headers = []
            for col_idx in range(1, ws.max_column + 1):
                cell = ws.cell(row=openpyxl_header_row, column=col_idx)
                headers.append(cell.value)
            
            st.write(f"ğŸ” [è°ƒè¯•] æ–‡ä»¶: {file_name}")
            st.write(f"ğŸ“‹ [è°ƒè¯•] åˆ—å: {headers}")
            
            # æ‰¾åˆ°æ˜µç§°åˆ—ç´¢å¼•
            nickname_col = None
            for idx, header in enumerate(headers, 1):
                if header:
                    for keyword in self.NICKNAME_KEYWORDS:
                        if keyword in str(header):
                            nickname_col = idx
                            break
                if nickname_col:
                    break
            
            st.write(f"ğŸ‘¤ [è°ƒè¯•] æ˜µç§°åˆ—ç´¢å¼•: {nickname_col}")
            
            if nickname_col is None:
                wb.close()
                st.warning(f"âš ï¸ [è°ƒè¯•] æœªæ‰¾åˆ°æ˜µç§°åˆ—ï¼Œè¿”å›ç©ºå­—å…¸")
                return {}
            
            # æ‰¾åˆ°æ‰€æœ‰å›¾ç‰‡ç›¸å…³çš„åˆ—ç´¢å¼•ï¼ˆæ’é™¤"è®¢æ­£å›¾ç‰‡"ï¼‰
            image_cols = []
            for idx, header in enumerate(headers, 1):
                if header and 'å›¾ç‰‡' in str(header):
                    # æ’é™¤è®¢æ­£å›¾ç‰‡ï¼Œåªç»Ÿè®¡ä¸Šä¼ çš„æˆªå›¾
                    if 'è®¢æ­£' not in str(header):
                        # æ£€æŸ¥æ˜¯å¦æ˜¯ç¼–å·çš„å›¾ç‰‡åˆ—ï¼ˆå›¾ç‰‡1, å›¾ç‰‡2...ï¼‰
                        header_str = str(header).strip()
                        if re.search(r'å›¾ç‰‡\d+', header_str):
                            image_cols.append(idx)
                            st.write(f"ğŸ–¼ï¸ [è°ƒè¯•] æ‰¾åˆ°å›¾ç‰‡åˆ—: {header_str} (ç´¢å¼•: {idx})")
            
            st.write(f"ğŸ“Š [è°ƒè¯•] å›¾ç‰‡åˆ—ç´¢å¼•æ€»æ•°: {len(image_cols)} ä¸ª")
            
            # ç»Ÿè®¡æ¯è¡Œçš„å›¾ç‰‡æ•°é‡
            nickname_image_count = {}
            for row_idx in range(openpyxl_header_row + 1, ws.max_row + 1):
                # è·å–æ˜µç§°
                nickname_cell = ws.cell(row=row_idx, column=nickname_col)
                nickname_raw = nickname_cell.value
                
                if not nickname_raw:
                    continue
                
                # æ¸…ç†æ˜µç§°
                nickname = self.clean_nickname(nickname_raw)
                if not nickname:
                    continue
                
                # ç»Ÿè®¡è¯¥è¡Œçš„å›¾ç‰‡æ•°é‡ï¼ˆåªç»Ÿè®¡æœ‰è¶…é“¾æ¥çš„ï¼‰
                image_count = 0
                for col_idx in image_cols:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    # åªæœ‰å½“å•å…ƒæ ¼æœ‰è¶…é“¾æ¥æ—¶ï¼Œæ‰ç®—ä½œæœ‰æ•ˆå›¾ç‰‡
                    if cell.hyperlink is not None:
                        image_count += 1
                
                # å¦‚æœæ˜µç§°å·²å­˜åœ¨ï¼Œå–æœ€å¤§å€¼ï¼ˆé¿å…é‡å¤æ¡ç›®ï¼‰
                if nickname in nickname_image_count:
                    nickname_image_count[nickname] = max(nickname_image_count[nickname], image_count)
                else:
                    nickname_image_count[nickname] = image_count
            
            # æ˜¾ç¤ºç»Ÿè®¡ç»“æœæ‘˜è¦
            total_people = len(nickname_image_count)
            total_images = sum(nickname_image_count.values())
            st.write(f"âœ… [è°ƒè¯•] ç»Ÿè®¡å®Œæˆ: å…± {total_people} äººï¼Œæ€»è®¡ {total_images} å¼ å›¾ç‰‡")
            
            # æ˜¾ç¤ºå‰5ä¸ªæ˜µç§°çš„ç»Ÿè®¡ï¼ˆç¤ºä¾‹ï¼‰
            if nickname_image_count:
                sample_items = list(nickname_image_count.items())[:5]
                st.write(f"ğŸ“ [è°ƒè¯•] ç¤ºä¾‹æ•°æ®: {dict(sample_items)}")
            
            wb.close()
            return nickname_image_count
            
        except Exception as e:
            # å¦‚æœç»Ÿè®¡å¤±è´¥ï¼Œè¿”å›ç©ºå­—å…¸ï¼ˆåç»­ä¼šä½¿ç”¨é»˜è®¤ç æ•°1ï¼‰
            st.error(f"âŒ [è°ƒè¯•] ç»Ÿè®¡å›¾ç‰‡æ•°é‡æ—¶å‡ºé”™: {str(e)}")
            import traceback
            st.code(traceback.format_exc())
            return {}
    
    def extract_nicknames_and_times_from_file(self, file_content, file_name: str) -> Tuple[List[str], List[str], List[int], str]:
        """
        ä»Excelæ–‡ä»¶å†…å®¹ä¸­æå–æ˜µç§°ã€æäº¤æ—¶é—´å’Œå›¾ç‰‡æ•°é‡ï¼ˆç æ•°ï¼‰
        
        Args:
            file_content: æ–‡ä»¶å†…å®¹ï¼ˆbytesæˆ–file-likeå¯¹è±¡ï¼‰
            file_name: æ–‡ä»¶å
            
        Returns:
            (æ˜µç§°åˆ—è¡¨, æäº¤æ—¶é—´åˆ—è¡¨, å›¾ç‰‡æ•°é‡åˆ—è¡¨, é”™è¯¯ä¿¡æ¯)
        """
        try:
            # é‡ç½®æ–‡ä»¶æŒ‡é’ˆï¼ˆå¦‚æœæ˜¯file-likeå¯¹è±¡ï¼‰
            if hasattr(file_content, 'seek'):
                file_content.seek(0)
            
            # å…ˆå°è¯•ç¬¬1è¡Œä¸ºåˆ—åï¼ˆheader=0ï¼‰
            header_row = 0
            if file_name.endswith('.xlsx'):
                df = pd.read_excel(file_content, engine='openpyxl', header=0)
            elif file_name.endswith('.xls'):
                df = pd.read_excel(file_content, engine='xlrd', header=0)
            else:
                return [], [], [], f"ä¸æ”¯æŒçš„æ–‡ä»¶æ ¼å¼: {file_name}"
            
            if df.empty:
                return [], [], [], f"æ–‡ä»¶ä¸ºç©º: {file_name}"
            
            # æŸ¥æ‰¾æ˜µç§°åˆ—
            nickname_column = self.find_nickname_column(df)
            
            # å¦‚æœç¬¬1è¡Œä½œä¸ºåˆ—åæ‰¾ä¸åˆ°æ˜µç§°åˆ—ï¼Œå°è¯•ç¬¬2è¡Œä½œä¸ºåˆ—åï¼ˆheader=1ï¼‰
            if nickname_column is None:
                try:
                    # é‡ç½®æ–‡ä»¶æŒ‡é’ˆå†æ¬¡è¯»å–
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    
                    header_row = 1
                    if file_name.endswith('.xlsx'):
                        df = pd.read_excel(file_content, engine='openpyxl', header=1)
                    elif file_name.endswith('.xls'):
                        df = pd.read_excel(file_content, engine='xlrd', header=1)
                    
                    if not df.empty:
                        nickname_column = self.find_nickname_column(df)
                except Exception as e:
                    pass
            
            # å¦‚æœè¿˜æ˜¯æ‰¾ä¸åˆ°æ˜µç§°åˆ—ï¼Œè¿”å›é”™è¯¯
            if nickname_column is None:
                # è¿”å›å¯ç”¨åˆ—åä¾›ç”¨æˆ·å‚è€ƒ
                available_columns = ", ".join(df.columns.tolist())
                return [], [], [], f"æœªæ‰¾åˆ°æ˜µç§°åˆ—ï¼Œå¯ç”¨åˆ—å: {available_columns}"
            
            # æŸ¥æ‰¾æ—¶é—´åˆ—
            time_column = self.find_time_column(df)
            
            # æå–æ˜µç§°å’Œæ—¶é—´æ•°æ®
            nicknames_raw = df[nickname_column].tolist()
            times_raw = df[time_column].tolist() if time_column else [None] * len(nicknames_raw)
            
            # ç»Ÿè®¡å›¾ç‰‡æ•°é‡ï¼ˆç æ•°ï¼‰
            nickname_image_count = self.count_images_in_excel(file_content, file_name, header_row)
            
            # æ¸…ç†å’Œå»é‡ï¼ŒåŒæ—¶ä¿æŒæ˜µç§°ã€æ—¶é—´å’Œå›¾ç‰‡æ•°é‡çš„å¯¹åº”å…³ç³»
            nicknames_clean = []
            times_clean = []
            image_counts = []
            seen = set()
            
            for nickname, time_val in zip(nicknames_raw, times_raw):
                clean_name = self.clean_nickname(nickname)
                if clean_name and clean_name not in seen and len(clean_name) > 0:
                    nicknames_clean.append(clean_name)
                    times_clean.append(str(time_val) if time_val is not None else "")
                    # è·å–è¯¥æ˜µç§°çš„å›¾ç‰‡æ•°é‡ï¼Œå¦‚æœæ²¡æœ‰åˆ™é»˜è®¤ä¸º1
                    image_count = nickname_image_count.get(clean_name, 1)
                    image_counts.append(image_count)
                    seen.add(clean_name)
            
            return nicknames_clean, times_clean, image_counts, ""
            
        except Exception as e:
            return [], [], [], f"å¤„ç†æ–‡ä»¶ {file_name} æ—¶å‡ºé”™: {str(e)}"
    
    def extract_nicknames_from_file(self, file_content, file_name: str) -> Tuple[List[str], str]:
        """
        ä»Excelæ–‡ä»¶å†…å®¹ä¸­æå–æ˜µç§°ï¼ˆä¿æŒå‘åå…¼å®¹ï¼‰
        
        Args:
            file_content: æ–‡ä»¶å†…å®¹ï¼ˆbytesï¼‰
            file_name: æ–‡ä»¶å
            
        Returns:
            (æ˜µç§°åˆ—è¡¨, é”™è¯¯ä¿¡æ¯)
        """
        nicknames, _, _, error = self.extract_nicknames_and_times_from_file(file_content, file_name)
        return nicknames, error
    
    def batch_process_files(self, uploaded_files) -> Tuple[List[Tuple[str, List[str]]], List[str]]:
        """
        æ‰¹é‡å¤„ç†ä¸Šä¼ çš„æ–‡ä»¶
        
        Args:
            uploaded_files: Streamlitä¸Šä¼ çš„æ–‡ä»¶åˆ—è¡¨
            
        Returns:
            (æˆåŠŸå¤„ç†çš„æ–‡ä»¶ç»“æœåˆ—è¡¨, é”™è¯¯ä¿¡æ¯åˆ—è¡¨)
        """
        successful_results = []
        error_messages = []
        
        for uploaded_file in uploaded_files:
            nicknames, error_msg = self.extract_nicknames_from_file(
                uploaded_file, uploaded_file.name
            )
            
            if error_msg:
                error_messages.append(f"{uploaded_file.name}: {error_msg}")
            else:
                successful_results.append((uploaded_file.name, nicknames))
        
        return successful_results, error_messages
    
    def validate_file_format(self, file_name: str) -> bool:
        """
        éªŒè¯æ–‡ä»¶æ ¼å¼æ˜¯å¦æ”¯æŒ
        
        Args:
            file_name: æ–‡ä»¶å
            
        Returns:
            æ˜¯å¦æ”¯æŒè¯¥æ ¼å¼
        """
        supported_extensions = ['.xlsx', '.xls']
        return any(file_name.lower().endswith(ext) for ext in supported_extensions)
    
    def get_file_info(self, file_content, file_name: str) -> dict:
        """
        è·å–Excelæ–‡ä»¶åŸºæœ¬ä¿¡æ¯
        
        Args:
            file_content: æ–‡ä»¶å†…å®¹ï¼ˆbytesæˆ–file-likeå¯¹è±¡ï¼‰
            file_name: æ–‡ä»¶å
            
        Returns:
            æ–‡ä»¶ä¿¡æ¯å­—å…¸
        """
        try:
            # é‡ç½®æ–‡ä»¶æŒ‡é’ˆï¼ˆå¦‚æœæ˜¯file-likeå¯¹è±¡ï¼‰
            if hasattr(file_content, 'seek'):
                file_content.seek(0)
            
            # å…ˆå°è¯•ç¬¬1è¡Œä¸ºåˆ—åï¼ˆheader=0ï¼‰
            if file_name.endswith('.xlsx'):
                df = pd.read_excel(file_content, engine='openpyxl', header=0)
            elif file_name.endswith('.xls'):
                df = pd.read_excel(file_content, engine='xlrd', header=0)
            else:
                return {"error": "ä¸æ”¯æŒçš„æ–‡ä»¶æ ¼å¼"}
            
            nickname_column = self.find_nickname_column(df)
            
            # å¦‚æœç¬¬1è¡Œæ‰¾ä¸åˆ°æ˜µç§°åˆ—ï¼Œå°è¯•ç¬¬2è¡Œä½œä¸ºåˆ—åï¼ˆheader=1ï¼‰
            if nickname_column is None:
                try:
                    # é‡ç½®æ–‡ä»¶æŒ‡é’ˆå†æ¬¡è¯»å–
                    if hasattr(file_content, 'seek'):
                        file_content.seek(0)
                    
                    if file_name.endswith('.xlsx'):
                        df = pd.read_excel(file_content, engine='openpyxl', header=1)
                    elif file_name.endswith('.xls'):
                        df = pd.read_excel(file_content, engine='xlrd', header=1)
                    nickname_column = self.find_nickname_column(df)
                except:
                    pass
            
            return {
                "file_name": file_name,
                "total_rows": len(df),
                "total_columns": len(df.columns),
                "columns": df.columns.tolist(),
                "nickname_column": nickname_column
            }
            
        except Exception as e:
            return {"error": f"è¯»å–æ–‡ä»¶ä¿¡æ¯å¤±è´¥: {str(e)}"}
